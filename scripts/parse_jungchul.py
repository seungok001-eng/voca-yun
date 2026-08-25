"""정철 교재 원본 엑셀 → 업로드용 JSON 변환.

원본은 레슨마다 시트가 하나씩인 세로 양식이다.
  T1~T8 = Toon World,  B1~B8 = Book Club
  1행: 교재/파트,  2행: "Lesson N 제목"
  [단어] : 번호|단어|뜻 (왼쪽 1~10) + 번호|단어|뜻 (오른쪽 11~20 = 심화)
  [문장] : "A : 영어 (한국어)" 줄이 이어지고, 빈 줄로 세트가 나뉜다
  [Story]: 본문 한 덩어리 (문장 단위로 잘라 저장한다)

Toon World는 기본용·심화용 두 파일로 나뉘어 오는데, 심화 파일이 기본을 포함하므로
두 파일을 함께 주면 기본 파일에 없는 단어를 '심화 전용'으로 표시한다.

사용:
  python3 scripts/parse_jungchul.py out.json \\
      --toon 1 기본.xlsx 심화.xlsx  --toon 2 기본2.xlsx 심화2.xlsx \\
      --book 1 북클럽1.xlsx          --book 2 북클럽2.xlsx
  (--toon/--book 뒤 숫자는 PART 번호. 심화 파일은 생략 가능)
"""
import json
import os
import re
import sys

import openpyxl

SKIP_SHEETS = {"가이드", "교재 로드맵", "샘플", "양식"}


def lesson_title(ws):
    """2행 'Lesson 1 Let's go to the circus now!' → (1, "Let's go to the circus now!")"""
    raw = str(ws.cell(2, 1).value or "").strip()
    m = re.match(r"Lesson\s*(\d+)\s*(.*)", raw, re.I)
    if not m:
        return None, raw
    return int(m.group(1)), m.group(2).strip()


def read_words(ws, col_no, col_text, col_mean):
    out = []
    for i in range(6, 30):
        text = ws.cell(i, col_text).value
        if not text:
            continue
        if not str(ws.cell(i, col_no).value or "").strip().isdigit():
            continue
        mean = str(ws.cell(i, col_mean).value or "").strip()
        out.append({"text": str(text).strip(), "meaning": mean})
    return out


def split_en_ko(raw):
    """"Let's go now! (우리 지금 가자!)" → ("Let's go now!", "우리 지금 가자!")"""
    s = str(raw).replace("\n", " ").strip()
    if not re.search(r"[가-힣]", s):
        return s, None
    # 한국어가 시작되는 첫 여는 괄호에서 자른다.
    # 원본에 닫는 괄호가 빠졌거나 "그의 형(남동생)이"처럼 괄호가 겹쳐 있어도 안전하다.
    m = re.search(r"\s*[（(]\s*(?=[^)）]*[가-힣])", s)
    if not m:
        return s, None
    en = s[:m.start()].strip()
    ko = s[m.end():].strip().rstrip(")）").strip()
    if not en or re.search(r"[가-힣]", en):
        return s, None
    return en, (ko or None)


def read_dialogues(ws):
    """[문장] 아래 A/B 줄을 빈 줄 기준으로 세트로 묶는다."""
    start = None
    for i in range(1, ws.max_row + 1):
        v = str(ws.cell(i, 1).value or "")
        if v.strip().startswith("[문장"):
            start = i + 1
            break
    if start is None:
        return []

    sets, cur = [], []
    for i in range(start, ws.max_row + 1):
        label = str(ws.cell(i, 1).value or "").strip()
        body = ws.cell(i, 2).value
        m = re.match(r"^([AB])\s*[:：]?$", label)
        if not m or not body:
            if cur:                      # 빈 줄 → 세트 종료
                sets.append(cur)
                cur = []
            continue
        text, ko = split_en_ko(body)
        if text:
            cur.append({"speaker": m.group(1), "text": text, **({"ko": ko} if ko else {})})
    if cur:
        sets.append(cur)
    return [{"lines": s} for s in sets if len(s) >= 2]


def read_passage(ws):
    """[Story] 본문을 문장 단위로 자른다."""
    start = None
    for i in range(1, ws.max_row + 1):
        if str(ws.cell(i, 1).value or "").strip().lower().startswith("[story"):
            start = i + 1
            break
    if start is None:
        return []

    chunks = []
    for i in range(start, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            v = ws.cell(i, j).value
            if v and len(str(v).strip()) > 20:
                chunks.append(str(v).strip())
    if not chunks:
        return []

    text = "\n\n".join(chunks)
    text = text.replace("’", "'").replace("“", '"').replace("”", '"')
    # 문장 끝(. ! ? 뒤 따옴표 포함) 다음 공백에서 자르되 Mr./Mrs. 같은 약어는 붙여둔다
    parts = re.split(r'(?<=[.!?])["\']?\s+', text.replace("\n\n", " "))
    out, buf = [], ""
    for p in parts:
        p = p.strip()
        if not p:
            continue
        buf = f"{buf} {p}".strip() if buf else p
        if re.search(r"\b(Mr|Mrs|Ms|Dr|St)\.$", buf):
            continue                      # 약어로 끝나면 다음 조각과 합친다
        out.append(buf)
        buf = ""
    if buf:
        out.append(buf)
    return [{"text": s} for s in out if len(s) > 1]


def parse_toon(part, basic_file, adv_file=None):
    """심화 파일이 있으면 기본 파일에 없는 단어를 심화 전용으로 표시한다."""
    basic_words = {}
    if basic_file:
        wb = openpyxl.load_workbook(basic_file, data_only=True)
        for ws in wb.worksheets:
            if ws.title in SKIP_SHEETS or not ws.title.upper().startswith("T"):
                continue
            n, _ = lesson_title(ws)
            if n:
                basic_words[n] = {w["text"].lower() for w in read_words(ws, 1, 2, 3)}

    src = adv_file or basic_file
    wb = openpyxl.load_workbook(src, data_only=True)
    lessons = []
    for ws in wb.worksheets:
        if ws.title in SKIP_SHEETS or not ws.title.upper().startswith("T"):
            continue
        n, title = lesson_title(ws)
        if not n:
            continue
        words = read_words(ws, 1, 2, 3) + read_words(ws, 5, 6, 7)
        known = basic_words.get(n)
        out_words = []
        for w in words:
            advanced = (known is not None) and (w["text"].lower() not in known)
            out_words.append({
                "text": w["text"],
                "meanings": [m.strip() for m in re.split(r"[,;/·]", w["meaning"]) if m.strip()] or [w["meaning"]],
                "advancedOnly": advanced,
            })
        lessons.append({
            "part": part, "area": "TOON", "order": n, "name": title,
            "words": out_words, "dialogues": read_dialogues(ws),
        })
    return lessons


def parse_book(part, path):
    wb = openpyxl.load_workbook(path, data_only=True)
    lessons = []
    for ws in wb.worksheets:
        if ws.title in SKIP_SHEETS or not ws.title.upper().startswith("B"):
            continue
        n, title = lesson_title(ws)
        if not n:
            continue
        words = read_words(ws, 1, 2, 3) + read_words(ws, 5, 6, 7)
        lessons.append({
            "part": part, "area": "READING", "order": n, "name": title,
            "words": [{
                "text": w["text"],
                "meanings": [m.strip() for m in re.split(r"[,;/·]", w["meaning"]) if m.strip()] or [w["meaning"]],
                "advancedOnly": False,
            } for w in words],
            "dialogues": read_dialogues(ws),
            "passageLines": read_passage(ws),
        })
    return lessons


# ── 폴더 통째로 읽기 ────────────────────────────────────────────
# 파일 첫 행("FEM PLANET 1-2 Part 1")과 시트 이름(T*/B*)으로 교재·파트·영역을 알아낸다.
# Toon World가 기본용·심화용 두 파일로 오면 단어가 많은 쪽을 심화 원본으로 삼는다.
def scan_dir(folder):
    import glob
    found = {}  # (교재, 파트, 영역) → [(단어수, 경로)]
    for f in sorted(glob.glob(os.path.join(folder, "*.xlsx"))):
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        except Exception:
            continue
        sheets = [w.title for w in wb.worksheets if w.title not in SKIP_SHEETS]
        if not sheets:
            continue
        ws = wb[sheets[0]]
        m = re.search(r"(PLANET|SKY)\s*(\d)-(\d)\s*Part\s*(\d)", str(ws.cell(1, 1).value or ""), re.I)
        if not m:
            print(f"  건너뜀(교재 확인 불가): {os.path.basename(f)}")
            continue
        book = f"{m.group(1).upper()} {m.group(2)}-{m.group(3)}"
        part = int(m.group(4))
        area = "TOON" if sheets[0].upper().startswith("T") else "READING"
        n = sum(len(read_words(wb[t], 1, 2, 3)) + len(read_words(wb[t], 5, 6, 7)) for t in sheets)
        found.setdefault((book, part, area), []).append((n, f))

    books = {}
    for (book, part, area), files in sorted(found.items()):
        files.sort()  # 단어가 적은 쪽 = 기본, 많은 쪽 = 심화
        basic = files[0][1]
        adv = files[-1][1] if len(files) > 1 else None
        lessons = parse_toon(part, basic, adv) if area == "TOON" else parse_book(part, basic)
        for L in lessons:
            if area == "TOON":
                L["isReview"] = L["order"] == 8            # Toon World L8은 복습
            elif book.startswith("SKY"):
                # SKY 북클럽은 복습 단원 없이 L1~4가 Book A, L5~8이 Book B
                L["bookLabel"] = "Book A" if L["order"] <= 4 else "Book B"
            else:
                L["isReview"] = L["order"] in (4, 8)       # PLANET 북클럽 L4·L8은 복습
        books.setdefault(book, []).extend(lessons)
        tag = "기본+심화" if adv else "기본만"
        print(f"  {book} P{part} {'Toon' if area == 'TOON' else 'Book'}: {len(lessons)}레슨 ({tag})")
    return books


def main():
    if sys.argv[1] == "--dir":
        folder, outdir = sys.argv[2], sys.argv[3]
        print("폴더 스캔:")
        books = scan_dir(folder)
        for book, lessons in sorted(books.items()):
            lessons.sort(key=lambda L: (L["part"], L["area"], L["order"]))
            slug = book.lower().replace(" ", "-")
            path = os.path.join(outdir, f"textbook-{slug}.json")
            json.dump({"lessons": lessons}, open(path, "w"), ensure_ascii=False, indent=1)
            w = sum(len(L.get("words", [])) for L in lessons)
            adv = sum(1 for L in lessons for x in L.get("words", []) if x.get("advancedOnly"))
            dl = sum(len(x["lines"]) for L in lessons for x in L.get("dialogues", []))
            pl = sum(len(L.get("passageLines", [])) for L in lessons)
            print(f"\n{book}: 레슨 {len(lessons)} · 단어 {w}(심화 {adv}) · 대화 {dl}문장 · 본문 {pl}문장 → {os.path.basename(path)}")
        return

    out_path = sys.argv[1]
    args = sys.argv[2:]
    lessons = []
    i = 0
    while i < len(args):
        kind = args[i]
        if kind == "--toon":
            part = int(args[i + 1])
            basic = args[i + 2]
            adv = args[i + 3] if i + 3 < len(args) and not args[i + 3].startswith("--") else None
            lessons += parse_toon(part, basic, adv)
            i += 4 if adv else 3
        elif kind == "--book":
            part = int(args[i + 1])
            lessons += parse_book(part, args[i + 2])
            i += 3
        else:
            i += 1

    lessons.sort(key=lambda L: (L["part"], L["area"], L["order"]))
    json.dump({"lessons": lessons}, open(out_path, "w"), ensure_ascii=False, indent=1)

    print(f"레슨 {len(lessons)}개 → {out_path}")
    for L in lessons:
        w = len(L.get("words", []))
        adv = sum(1 for x in L.get("words", []) if x.get("advancedOnly"))
        d = len(L.get("dialogues", []))
        dl = sum(len(x["lines"]) for x in L.get("dialogues", []))
        p = len(L.get("passageLines", []))
        area = "Toon" if L["area"] == "TOON" else "Book"
        print(f"  P{L['part']} {area} L{L['order']:<2} 단어 {w:>2}(심화 {adv}) · 대화 {d}세트({dl}문장) · 본문 {p}문장 · {L['name'][:34]}")


if __name__ == "__main__":
    main()
